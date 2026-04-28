# /// script
# requires-python = ">=3.11"
# dependencies = [
#     "openpyxl>=3.1",
# ]
# ///

"""Read the audit xlsx and report user's decisions"""

from pathlib import Path

from openpyxl import load_workbook

xlsx_path = Path.home() / "hermes-workspace" / "hermes-skills" / "技能审计报告.xlsx"
wb = load_workbook(xlsx_path)
ws = wb.active

# Headers: 技能名称 | 来源 | 启用状态 | 描述 | 审计建议 | 我的决策
changes = []
for row in ws.iter_rows(min_row=2, values_only=True):
    name, source, enabled, desc, suggestion, decision = row
    if not name:
        continue
    decision = str(decision).strip() if decision else ""

    # Determine what the default would be
    is_builtin = source == "内置"
    if is_builtin:
        default = "启用" if enabled == "启用" else "禁用"
    else:
        default = "保留"

    if decision != default:
        changes.append(
            {
                "name": name,
                "source": source,
                "enabled": enabled,
                "suggestion": suggestion,
                "default": default,
                "decision": decision,
            }
        )

# Categorize
delete_list = [c for c in changes if c["decision"] == "删除"]
disable_list = [c for c in changes if c["decision"] == "禁用"]
enable_list = [c for c in changes if "启用" in c["decision"] and c["decision"] != c["default"]]
other = [c for c in changes if c["decision"] not in ("删除", "禁用", "启用") and c["decision"]]

print("=== 技能审计决策汇总 ===")
print(f"总技能数: {ws.max_row - 1}")
print(f"有变更的技能: {len(changes)}")
print()

if delete_list:
    print(f"--- 标记删除 ({len(delete_list)} 个) ---")
    for c in sorted(delete_list, key=lambda x: x["name"]):
        print(f"  🗑️  {c['name']} ({c['source']})")
    print()

if disable_list:
    print(f"--- 标记禁用 ({len(disable_list)} 个) ---")
    for c in sorted(disable_list, key=lambda x: x["name"]):
        print(f"  ⚠️  {c['name']} (内置)")
    print()

if enable_list:
    print(f"--- 标记启用 ({len(enable_list)} 个) ---")
    for c in sorted(enable_list, key=lambda x: x["name"]):
        print(f"  ✅  {c['name']} (内置)")
    print()

if other:
    print(f"--- 其他变更 ({len(other)} 个) ---")
    for c in sorted(other, key=lambda x: x["name"]):
        print(f"  ❓ {c['name']}: {c['decision']}")
    print()

print("================")
