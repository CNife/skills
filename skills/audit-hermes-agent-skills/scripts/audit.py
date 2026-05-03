# /// script
# requires-python = ">=3.11"
# dependencies = [
#     "pyyaml>=6.0",
#     "openpyxl>=3.1",
# ]
# ///

"""Hermes Agent 技能审计 — 生成 XLSX 并执行清理决策

用法：
    uv run audit.py             扫描技能 → 打印概览 → 生成 XLSX
    uv run audit.py --apply     从 XLSX 读取决策 → 打印变更 → 确认后执行
"""

import argparse
import json
import shutil
import sqlite3
import subprocess
from datetime import datetime
from pathlib import Path

import yaml
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── 路径 ────────────────────────────────────────────────────────────────────
HERMES_HOME = Path.home() / ".hermes"
SKILLS_DIR = HERMES_HOME / "skills"
AGENTS_SKILLS = Path.home() / ".agents" / "skills"
STATE_DB = HERMES_HOME / "state.db"
CONFIG_PATH = HERMES_HOME / "config.yaml"
_EXCLUDED = frozenset((".git", ".github", ".hub", ".audit-backups"))

# ── 中文描述映射 ─────────────────────────────────────────────────────────────
CN_DESCRIPTIONS = {
    "agent-browser": "AI Agent 浏览器自动化工具，支持页面导航、表单填写、内容提取等操作",
    "airtable": "通过 curl 调用 Airtable REST API，管理记录增删改查与过滤",
    "apple-notes": "通过 memo CLI 管理 Apple 备忘录：创建、搜索、编辑",
    "apple-reminders": "通过 remindctl 管理 Apple 提醒事项：添加、列表、完成",
    "architecture-diagram": "深色主题 SVG 架构/云/基础设施图，输出为 HTML",
    "arxiv": "按关键词、作者、分类或 ID 搜索 arXiv 论文",
    "ascii-art": "ASCII 艺术：pyfiglet、cowsay、boxes、图片转字符画",
    "ascii-video": "ASCII 视频：将视频/音频转换为彩色 ASCII MP4/GIF",
    "audiocraft": "AudioCraft：MusicGen 文生音乐、AudioGen 文生音效",
    "audiocraft-audio-generation": "AudioCraft：MusicGen 文生音乐、AudioGen 文生音效",
    "axolotl": "Axolotl：通过 YAML 配置进行 LoRA/DPO/GRPO 等 LLM 微调",
    "backtest-expert": "量化回测策略开发与验证专家，覆盖开发、测试、压力测试与归因分析",
    "blogwatcher": "通过 blogwatcher-cli 监控博客和 RSS/Atom 订阅源",
    "camofox-browser": "基于 Camoufox 的反检测无头浏览器自动化服务，绕过 Google 等网站的机器人检测",
    "check": "实现后审查代码 diff，自动修复安全问题，并针对大 diff 运行安全与架构审查",
    "chezmoi-workflows": "使用 chezmoi 管理 dotfiles 备份与同步",
    "clash-verge-rev": "从 WSL 管理 Clash Verge Rev 代理客户端，修改 TUN 模式、代理设置等配置并重启",
    "claude-code": "委托编码任务给 Claude Code CLI（实现功能、创建 PR）",
    "claude-design": "设计一次性 HTML 作品（落地页、演示文稿、原型）",
    "clawdchat": "ClawdChat AI Agent 社交网络与通用工具网关，支持帖子、评论、投票、私信、群组等功能",
    "clawdchat-onboarding": "注册并入驻 ClawdChat.ai AI 社交网络账号，包括注册、凭证提取、认领验证等流程",
    "clawdchat-registration": "注册 ClawdChat.ai 账号，处理 API Key 提取和凭证保存",
    "clip": "OpenAI 视觉+语言连接模型，支持零样本图像分类和图文匹配",
    "cloud-to-internal-migration": "将云托管服务（数据库、对象存储、应用）迁移至内部基础设施的系统化流程",
    "cnife-skills-repo": "CNife/skills 仓库结构指南，包含 Python 项目规范、pre-commit hooks 和 PEP 723 脚本规范",
    "codebase-inspection": "使用 pygount 检查代码库：代码行数、语言分布、比例",
    "codex": "委托编码任务给 OpenAI Codex CLI（实现功能、创建 PR）",
    "creative-ideation": "通过创意约束生成项目点子",
    "debugging-hermes-tui-commands": "调试 Hermes TUI 斜杠命令：Python、Gateway、Ink UI",
    "design": "为任意组件、页面或界面生成独特的高质量 UI，支持截图驱动迭代",
    "design-md": "编写/验证/导出 Google DESIGN.md 令牌规范文件",
    "dida365-creation-helper": "标准工作流用于创建滴答清单任务，解决换行符泄露、默认列表错误和时间语义问题",
    "dida365-openapi": "基于滴答清单官方 OpenAPI 和 OAuth2 的任务管理 Skill，支持任务的创建/修改/完成/查询",
    "dogfood": "探索性 QA 测试网页应用：发现 Bug、收集证据、生成报告",
    "dspy": "DSPy：声明式语言模型编程，自动优化提示词和检索增强生成",
    "evaluating-llms-harness": "lm-eval-harness：评估 LLM 性能（MMLU、GSM8K 等基准）",
    "excalidraw": "手绘风格 Excalidraw JSON 图表（架构图、流程图、时序图）",
    "feishu-cli-usage": "飞书/Lark CLI 使用指南，专注消息类型、模式和集成",
    "find-nearby": "通过 OpenStreetMap 查找附近地点（餐厅、咖啡馆、酒吧、药店等）",
    "find-skills": "帮助用户发现和安装 Agent 技能，当用户询问「如何做某事」时使用",
    "findmy": "通过 macOS FindMy.app 跟踪 Apple 设备/AirTag",
    "fine-tuning-with-trl": "TRL：SFT、DPO、PPO、GRPO、Reward Modeling 等 LLM 强化学习微调",
    "frontend-design": "创建具有高设计质量的前端界面，用于页面设计、组件构建和交互体验优化",
    "gguf": "GGUF 格式与 llama.cpp 量化，用于高效的 CPU/GPU 推理部署",
    "gguf-quantization": "GGUF 格式与 llama.cpp 量化，用于高效的 CPU/GPU 推理部署",
    "gif-search": "通过 curl + jq 从 Tenor 搜索和下载 GIF",
    "git-master": "所有 Git 操作的必备技能，包含原子提交、rebase/squash、历史搜索（blame、bisect、log -S）",
    "github-auth": "GitHub 认证配置：HTTPS Token、SSH Key、gh CLI 登录",
    "github-code-review": "审查 PR：通过 gh 或 REST API 查看 diff 和内联评论",
    "github-issues": "通过 gh 或 REST API 创建、分类、标记、分配 GitHub Issue",
    "github-pr-workflow": "GitHub PR 生命周期：分支、提交、打开、CI、合并",
    "github-repo-management": "克隆/创建/复刻仓库，管理远程仓库和发布",
    "github-trending-cli": "在终端查看 GitHub Trending 热门仓库",
    "gitlab-workflow": "GitLab 最佳实践：Merge Request、CI/CD 流水线、Issue 跟踪和 DevOps 工作流",
    "godmode": "LLM 越狱：Parseltongue、GODMODE、ULTRAPLINIAN",
    "google-workspace": "通过 gws CLI 或 Python 访问 Gmail、日历、Drive、文档和表格",
    "grpo-rl-training": "GRPO/RL 微调专家指南，使用 TRL 进行推理和任务特定模型训练",
    "guidance": "使用正则和文法控制 LLM 输出，保证生成有效的 JSON/XML/代码结构",
    "heartmula": "HeartMuLa：类似 Suno 的基于歌词+标签的歌曲生成",
    "hermes-agent": "Hermes Agent 使用与扩展完整指南：CLI 用法、配置、网关、工具、技能和功能",
    "hermes-agent-daily-changelog": "自动抓取并总结 Hermes Agent 仓库最近 25 小时的 Git 提交变更",
    "hermes-agent-skill-authoring": "编写仓库内联 SKILL.md：前置元数据、验证器、目录结构",
    "hermes-file-structure": "Hermes Agent 文件结构指南：配置、记忆、人格文件位置",
    "hermes-media-delivery": "Hermes Agent 中跨渠道正确发送媒体的指南",
    "hermes-skills": "Hermes Agent 技能安装、更新和管理指南",
    "hermes-token-analysis": "通过本地代理分析 Hermes Agent API Token 消耗明细",
    "hermes-usage-insights": "查询 Hermes Agent 汇总使用统计：Token、缓存命中率、费用趋势",
    "hermes-webui-baremetal": "裸金属 systemd 部署 hermes-webui，涵盖 bootstrap.py fork 陷阱",
    "hermes-webui-podman": "Podman 部署 hermes-webui，含容器启动参数和 systemd 集成",
    "himalaya": "通过 IMAP/SMTP 管理邮件的 CLI 工具，支持列表、阅读、编写、回复、转发、搜索",
    "hindsight-local-embedded-setup": "Hermes Agent 本地嵌入模式 Hindsight 记忆系统排错与配置",
    "hindsight-memory-setup": "配置和排错 Hermes Agent 本地嵌入模式的 Hindsight 记忆系统",
    "huggingface-hub": "HuggingFace hf CLI：搜索/下载/上传模型和数据集",
    "hunt": "在应用修复前定位错误、崩溃、异常行为和失败测试的根因",
    "ideation": "通过创意约束生成项目点子",
    "imessage": "通过 macOS 的 imsg CLI 发送和接收 iMessage/SMS",
    "iwencai-skillhub": "安装和配置 Iwencai SkillHub CLI 及技能，涵盖手动安装和依赖管理",
    "json-canvas": "创建和编辑 JSON Canvas 文件（.canvas），支持节点、边、分组和连接线",
    "jupyter-live-kernel": "通过实时 Jupyter 内核（hamelnb）进行迭代式 Python 开发",
    "karpathy-guidelines": "减少常见 LLM 编码错误的行为指南，在编写、审查或重构代码时使用",
    "lab-check-api-reverse": "逆向工程的 Chingo lab4j-server 平台 API 客户端（之江实验室及 20+ 高校使用）",
    "lab-safe-check": "使用 Playwright 浏览器自动化，自动完成之江实验室专业实验室的每日安全自查",
    "learn": "运行六阶段研究工作流，将陌生领域或收集的素材转化为可发布的输出",
    "linear": "通过 GraphQL + curl 管理 Linear 项目：Issue、团队、项目",
    "llama-cpp": "llama.cpp 本地 GGUF 推理 + HF Hub 模型发现",
    "llama-cpp-performance-tuning": "llama.cpp GPU 推理性能调优，包含自动化基准测试脚本和参数优化",
    "lm-evaluation-harness": "lm-eval-harness：评估 LLM 性能（MMLU、GSM8K 等基准）",
    "mcp-deepwiki": "访问和搜索 DeepWiki/GitHub 公共代码仓库文档的技能",
    "mcporter": "使用 mcporter CLI 列出、配置、认证和调用 MCP 服务器/工具（HTTP 或 stdio）",
    "mermaid-diagrams": "使用 Mermaid 语法创建软件系统图表的综合指南",
    "minecraft-modpack-server": "托管模组版 Minecraft 服务器（CurseForge、Modrinth）",
    "modal": "无服务器 GPU 云平台，按需 GPU 算力无需管理基础设施",
    "modal-serverless-gpu": "无服务器 GPU 云平台，按需 GPU 算力无需管理基础设施",
    "nano-pdf": "通过 nano-pdf CLI 编辑 PDF 文本/错字/标题",
    "native-mcp": "MCP 客户端：连接服务器、注册工具（stdio/HTTP）",
    "neverland-farm": "Neverland 农场自动化维护 — 技能驱动的自进化农场管理系统",
    "node-inspect-debugger": "通过 --inspect + Chrome DevTools Protocol CLI 调试 Node.js",
    "notion": "通过 curl 调用 Notion API：页面、数据库、区块、搜索",
    "obliteratus": "OBLITERATUS：通过 diff-in-means 消除 LLM 拒绝回答行为",
    "obsidian": "在 Obsidian 知识库中阅读、搜索和创建笔记",
    "obsidian-bases": "创建和编辑 Obsidian Bases（.base 文件），支持视图、筛选器、公式和汇总",
    "obsidian-cli": "使用 Obsidian CLI 与知识库交互，支持阅读、创建、搜索和管理笔记",
    "obsidian-diary": "将会话内容总结到 Obsidian 工作日志或个人日记中",
    "obsidian-markdown": "创建和编辑 Obsidian 风味 Markdown，支持 Wiki 链接、嵌入、Callout 和属性",
    "ocr-and-documents": "从 PDF/扫描件中提取文本（pymupdf、marker-pdf）",
    "opencli-adapter-author": "为新站点编写 OpenCLI 适配器或为已有站点添加新命令",
    "opencli-autofix": "自动修复 OpenCLI 适配器故障",
    "opencli-browser": "通过 opencli 驱动真实 Chrome 浏览器窗口",
    "opencli-usage": "OpenCLI 顶层使用指南，包含适配器发现、命令查找和工作流模式",
    "opencode": "委托编码任务给 OpenCode CLI（实现功能、PR 审查）",
    "opencode-primitives": "编写技能、插件、MCP 或配置驱动行为时参考 OpenCode 文档",
    "openhue": "通过 OpenHue CLI 控制 Philips Hue 灯光、场景和房间",
    "outlines": "Outlines：结构化 JSON/正则/Pydantic LLM 生成",
    "p5js": "p5.js 作品：生成艺术、着色器、交互、3D 图形",
    "paseo-install": "从源码在 WSL/Linux 上安装和配置 Paseo AI Agent 管理守护进程",
    "pdf": "PDF 文件处理技能：阅读、提取文本/表格、编辑、转换和生成 PDF",
    "peft": "使用 LoRA、QLoRA 等 25+ 方法进行参数高效微调",
    "peft-fine-tuning": "使用 LoRA、QLoRA 等 25+ 方法进行参数高效微调",
    "pixel-art": "像素艺术，支持 NES、Game Boy、PICO-8 等时代调色板",
    "plan": "计划模式：将 Markdown 计划写入 .hermes/plans/，不执行代码",
    "playwright-best-practices": "Playwright 测试编写、修复不稳定测试、调试失败指南",
    "pokemon-player": "通过无头模拟器 + 内存读取游玩宝可梦",
    "polymarket": "查询 Polymarket：市场、价格、订单簿、历史数据",
    "popular-web-designs": "54 个真实设计系统（Stripe、Linear、Vercel）的 HTML/CSS 实现",
    "powerpoint": "创建、读取、编辑 .pptx 演示文稿、幻灯片、备注和模板",
    "product-manager-toolkit": "现代产品管理的核心工具和框架，从需求发现到交付",
    "python-debugpy": "Python 调试：pdb REPL + debugpy 远程调试（DAP 协议）",
    "pytorch-fsdp": "PyTorch FSDP 全分片数据并行训练专家指南",
    "quantitative-backtest": "基于 Tushare 数据的 ETF/股票量化回测流程",
    "quantitative-research": "世界级系统化交易研究：回测、Alpha 生成、因子模型、统计套利",
    "read": "将任意 URL 或 PDF 转换为干净 Markdown，支持付费墙、JS 重页面和中文平台",
    "requesting-code-review": "提交前审查：安全扫描、质量门禁、自动修复",
    "research-paper-writing": "撰写 NeurIPS/ICML/ICLR 级别的机器学习论文",
    "segment-anything": "SAM：通过点、框、掩码进行零样本图像分割",
    "segment-anything-model": "SAM：通过点、框、掩码进行零样本图像分割",
    "service-guard": "跨平台服务监控，监控 WSL systemd 和 Windows 原生服务，检测隧道连接问题",
    "serving-llms-vllm": "vLLM：高吞吐 LLM 服务部署，OpenAI API 兼容，支持量化",
    "skill-creator": "创建、修改和改进技能，并评估技能性能",
    "skill-evaluator": "评估、比较、推荐、发现和安装 AI Agent 技能",
    "skill-hermes": "去除 LLM 废话的系统提示词，减少 56-73% 的啰嗦输出",
    "smart-search": "基于 opencli 命令的智能搜索路由器，用于在指定网站搜索信息",
    "software-design-philosophy": "基于 John Ousterhout《软件设计哲学》的软件设计指南",
    "songsee": "音频频谱图/特征提取（Mel、色度、MFCC）",
    "songwriting-and-ai-music": "歌曲创作技巧与 Suno AI 音乐提示词",
    "spotify": "Spotify：播放、搜索、排队、管理播放列表和设备",
    "stable-diffusion": "通过 HuggingFace Diffusers 使用 Stable Diffusion 模型进行文生图",
    "stable-diffusion-image-generation": "通过 HuggingFace Diffusers 使用 Stable Diffusion 模型进行文生图",
    "storage-cleanup": "扫描系统磁盘空间，识别浪费空间并清理",
    "subagent-driven-development": "通过 delegate_task 子代理执行计划（两阶段审查）",
    "systematic-debugging": "四阶段根因调试：理解 Bug 后再修复",
    "talk-normal": "去除 LLM 废话的系统提示词，显著减少啰嗦输出同时保留有用信息",
    "technical-svg-diagrams": "生成 Cloudflare 风格的简洁技术 SVG 图表",
    "test-driven-development": "TDD：强制 RED-GREEN-REFACTOR 流程，测试先于代码",
    "think": "将粗略想法转化为经验证的计划，编写代码前先确认结构",
    "tushare": "面向中文自然语言的 Tushare 数据研究技能",
    "tushare-data": "面向中文自然语言的 Tushare 数据研究技能",
    "tushare-finance": "获取中国金融市场数据，支持 220+ Tushare Pro 接口",
    "using-git-worktrees": "当功能开发需要与当前工作区隔离时使用 Git Worktree",
    "uv-package-manager": "使用 uv 包管理器进行快速 Python 依赖管理和虚拟环境",
    "vllm": "vLLM：高吞吐 LLM 服务部署，OpenAI API 兼容，支持量化",
    "web-access": "所有联网操作处理入口，包括搜索、网页抓取、登录后操作和网络交互",
    "webhook-subscriptions": "Webhook 订阅：事件驱动的 Agent 自动运行",
    "weights-and-biases": "W&B：记录 ML 实验、超参数搜索、模型注册表和仪表盘",
    "whisper": "OpenAI 通用语音识别模型，支持 99 种语言的转录、翻译和语言识别",
    "windows-browser-control": "从 WSL 通过 Chrome DevTools Protocol 控制 Windows 浏览器",
    "write": "去除 AI 写作痕迹，让文章读起来像自然的中文或英文",
    "writing-plans": "编写实现计划：小任务拆分、文件路径、代码结构",
    "wsl-memory-diagnostics": "WSL 内存问题诊断工作流，区分 WSL 内部使用与 Windows 主机消耗",
    "wsl-service-watchdog": "已废弃，由 service-guard 替代",
    "xitter": "通过 x-cli 终端客户端使用官方 X API 与 Twitter/X 交互",
    "xurl": "通过 xurl CLI 使用 X/Twitter v2 API：发帖、搜索、私信、媒体",
    "youtube-content": "将 YouTube 转录转换为摘要、推文、博客文章",
    "zjlab-model-deploy": "在 zjlab.icp-dev-1 上部署 LLM 模型（GGUF），支持服务启停、镜像下载和 vLLM 配置",
    "audit-hermes-agent-skills": "审计 Hermes Agent 已安装技能，统计使用频率并生成清理建议",
    "arch-wsl-cleanup": "通过分层清理策略释放 Arch Linux WSL 磁盘空间",
    "arch-wsl-install": "通过 pacman、AUR 或其他方式在 Arch Linux WSL 上安装包",
    "a-share-value-analysis": "A股价值投资分析工具，使用 Tushare 获取数据，进行 PE/PB 历史分位、DCF 估值等分析",
    "cronjob-feishu-fix": "修复 Hermes cron job 飞书消息投递失败的问题",
    "diagnose-slow-session-search": "诊断 Hermes Agent session_search 性能缓慢的根因",
    "doubao-share-extract": "提取豆包分享链接的完整对话内容并整理为文章，保存到 Obsidian 个人库",
    "entrocamp": "逆熵进化营 Agent 能力评估与考试系统",
    "github-trending": "抓取 GitHub Trending 热门仓库并格式化输出",
    "hermes-gateway-troubleshoot": "诊断和修复 Hermes gateway 网关服务故障",
    "optimize-agents-md": "AGENTS.md 编写与优化指南，遵循渐进式披露原则",
    "llm-wiki": "构建和查询互联的 Markdown 知识库",
    "manim-video": "Manim CE 动画：3Blue1Brown 风格的数学/算法视频",
    "maps": "通过 OpenStreetMap/OSRM 进行地理编码、POI、路线规划",
    "trf-fine-tuning": "TRL 微调：SFT、DPO、PPO、GRPO 和 Reward Modeling",
    "unsloth": "Unsloth：2-5 倍更快的 LoRA/QLoRA 微调，更少显存消耗",
}


# ── Hermes 内部 API ─────────────────────────────────────────────────────────
def find_hermes_venv_python() -> Path | None:
    hermes_bin = shutil.which("hermes")
    if not hermes_bin:
        return None
    try:
        with open(hermes_bin) as f:
            first_line = f.readline().strip()
            if first_line.startswith("#!"):
                python_path = first_line[2:].strip()
                if Path(python_path).exists():
                    return Path(python_path)
    except (OSError, PermissionError):
        pass
    return None


def get_skill_registry() -> dict[str, dict] | None:
    python = find_hermes_venv_python()
    if not python:
        return None
    try:
        result = subprocess.run(
            [
                str(python),
                "-c",
                r"""
import json
from tools.skills_tool import _find_all_skills
from tools.skills_sync import _read_manifest
from tools.skills_hub import HubLockFile

all_skills = _find_all_skills(skip_disabled=True)
hub_names = {e['name'] for e in HubLockFile().list_installed()}
builtin_names = set(_read_manifest())

output = {}
for s in all_skills:
    name = s['name']
    if name in hub_names:
        source = 'hub'
    elif name in builtin_names:
        source = 'builtin'
    else:
        source = 'local'
    output[name] = {'source': source, 'category': s.get('category')}
print(json.dumps(output))
""",
            ],
            capture_output=True,
            text=True,
            timeout=15,
        )
        if result.returncode == 0 and result.stdout.strip():
            return json.loads(result.stdout)
    except (subprocess.TimeoutExpired, json.JSONDecodeError, OSError):
        pass
    return None


def get_disabled_set() -> set[str]:
    if not CONFIG_PATH.exists():
        return set()
    try:
        with open(CONFIG_PATH) as f:
            config = yaml.safe_load(f) or {}
        return set(config.get("skills", {}).get("disabled", []))
    except Exception:
        return set()


# ── Curator 数据集成 ──────────────────────────────────────────────────────────
def parse_curator_activity(line: str) -> dict | None:
    """解析 curator status 单行输出为活跃度字典。"""
    parts = line.strip().split()
    if len(parts) < 2:
        return None
    activity = {}
    for part in parts[1:]:
        if "=" in part:
            k, v = part.split("=", 1)
            activity[k] = v
    return activity if activity else None


def get_curator_data() -> dict:
    """从 Hermes Curator 获取技能来源分类和活跃度数据。

    Returns:
        {
            "agent_created": set[str],     # curator 认定的首次方技能
            "activity": dict[str, dict],   # name -> {activity, use, view, patches, last_activity}
            "consolidated": dict[str, str],  # name -> umbrella_name
            "archived": set[str],           # 已归档技能
        }
    """
    result: dict = {
        "agent_created": set(),
        "activity": {},
        "consolidated": {},
        "archived": set(),
    }

    # Method 1: Parse hermes curator status
    try:
        proc = subprocess.run(
            ["hermes", "curator", "status"],
            capture_output=True, text=True, timeout=10,
        )
        if proc.returncode == 0:
            # Scan all lines for skill rows matching: name  activity=N  use=N ...
            # The curator status has 3 skill tables (least active / most active / least recently
            # active), plus summary lines (active/stale/archived counts). We skip any line where
            # the first token is a section header or a count label.
            skip_first = {"agent", "active", "stale", "archived", "least", "most", "curator:"}
            lines = proc.stdout.splitlines()
            for line in lines:
                stripped = line.strip()
                if not stripped:
                    continue
                parts = stripped.split()
                if len(parts) < 2:
                    continue
                first = parts[0].rstrip(":")
                if first in skip_first:
                    continue
                # Check if second token looks like "activity=N"
                if not parts[1].startswith("activity="):
                    continue
                # This is a skill row
                name = parts[0]
                result["agent_created"].add(name)
                act = parse_curator_activity(stripped)
                if act:
                    result["activity"][name] = act
    except (subprocess.TimeoutExpired, OSError, FileNotFoundError):
        pass

    # Method 2: Read latest run.json for consolidation and archive data
    try:
        import glob
        log_dirs = sorted(
            glob.glob(str(HERMES_HOME / "logs" / "curator" / "*/")), reverse=True
        )
        if log_dirs:
            run_path = Path(log_dirs[0]) / "run.json"
            if run_path.exists():
                run_data = json.loads(run_path.read_text(encoding="utf-8"))
                for c in run_data.get("consolidated", []):
                    result["consolidated"][c["name"]] = c["into"]
                for a in run_data.get("archived", []):
                    result["archived"].add(a)
    except (OSError, json.JSONDecodeError, IndexError):
        pass

    return result


# ── 技能仓库来源分类 ──────────────────────────────────────────────────────────
def get_skill_source_manifest() -> dict[str, str]:
    """从锁文件精确识别技能是通过 skills.sh 还是 SkillHub 安装的。

    Returns:
        {skill_name: "skills.sh"} | {skill_name: "skillhub"}
    """
    manifest: dict[str, str] = {}
    agents_lock = Path.home() / ".agents" / ".skill-lock.json"
    if agents_lock.exists():
        try:
            data = json.loads(agents_lock.read_text(encoding="utf-8"))
            for name in data.get("skills", {}):
                manifest[name] = "skills.sh"
        except (OSError, json.JSONDecodeError):
            pass
    hub_lock = SKILLS_DIR / ".hub" / "lock.json"
    if hub_lock.exists():
        try:
            data = json.loads(hub_lock.read_text(encoding="utf-8"))
            for name in data.get("installed", {}):
                manifest[name] = "skillhub"
        except (OSError, json.JSONDecodeError):
            pass
    return manifest


# ── 技能扫描 ─────────────────────────────────────────────────────────────────
def scan_all_skills() -> dict[str, dict]:
    skills = {}

    def scan_directory(base_dir: Path, is_external: bool = False):
        if not base_dir.exists():
            return
        for skill_md in base_dir.rglob("SKILL.md"):
            if any(part in _EXCLUDED for part in skill_md.parts):
                continue
            name = skill_md.parent.name
            if name in skills:
                continue
            parent = skill_md.parent.parent
            category = parent.name if parent != base_dir else None
            skills[name] = {
                "dir": str(skill_md.parent),
                "category": category,
                "is_external": is_external,
            }

    scan_directory(SKILLS_DIR)
    scan_directory(AGENTS_SKILLS, is_external=True)

    registry = get_skill_registry()
    disabled = get_disabled_set()
    curator_data = get_curator_data()
    source_manifest = get_skill_source_manifest()
    for name, info in skills.items():
        info["disabled"] = name in disabled
        if registry and name in registry:
            info["source"] = registry[name]["source"]
            if registry[name].get("category") and not info["category"]:
                info["category"] = registry[name]["category"]
        else:
            info["source"] = "external" if info["is_external"] else "local"

        # 覆盖 1: Curator agent-created 比 local 更精确
        if name in curator_data["agent_created"]:
            info["source"] = "agent-created"

        # 覆盖 2: 锁文件来源比 API 的 "hub" 更精确（区分 skills.sh vs skillhub）
        if name in source_manifest:
            info["source"] = source_manifest[name]

        # Curator 活跃度数据
        if name in curator_data.get("activity", {}):
            info["curator_activity"] = curator_data["activity"][name]

        # Consolidation / Archive 感知
        if name in curator_data.get("consolidated", {}):
            info["consolidated_into"] = curator_data["consolidated"][name]
        if name in curator_data.get("archived", {}):
            info["archived"] = True

        info["description"] = CN_DESCRIPTIONS.get(name, "")
    return skills


def count_skill_calls() -> dict[str, dict]:
    if not STATE_DB.exists():
        return {}
    conn = sqlite3.connect(str(STATE_DB))
    cur = conn.cursor()
    cur.execute("""
        SELECT tool_calls, timestamp FROM messages
        WHERE tool_calls IS NOT NULL
        AND (tool_calls LIKE '%skill_view%' OR tool_calls LIKE '%skill_manage%')
    """)
    stats = {}
    for tool_calls_json, timestamp in cur.fetchall():
        try:
            items = json.loads(tool_calls_json)
            for item in items:
                fn = item.get("function", {}).get("name", "")
                if fn in ("skill_view", "skill_manage"):
                    args = json.loads(item.get("function", {}).get("arguments", "{}"))
                    name = args.get("name", "")
                    if not name:
                        continue
                    if name not in stats:
                        stats[name] = {"total": 0, "last_seen": 0}
                    stats[name]["total"] += 1
                    if timestamp > stats[name]["last_seen"]:
                        stats[name]["last_seen"] = timestamp
        except (json.JSONDecodeError, KeyError):
            continue
    conn.close()
    return stats


# ── XLSX 生成 ────────────────────────────────────────────────────────────────
def format_source(source: str) -> str:
    return {
        "builtin": "内置",
        "skills.sh": "skills.sh安装",
        "skillhub": "SkillHub安装",
        "agent-created": "Agent创建",
        "local": "本地创建",
        "hub": "skills.sh安装",
        "external": "外部共享",
    }.get(source, source)


def get_suggestion(name: str, info: dict, stats: dict | None) -> str:
    # 已归档技能 — 最高优先级提示
    if info.get("archived"):
        return "📦 已归档，建议删除（已被 curator 归档）"
    # 已被合并到 umbrella
    if info.get("consolidated_into"):
        return f"🔄 已合并到 {info['consolidated_into']}，建议删除原始技能"
    if info.get("disabled"):
        return "已禁用，无需操作"
    if not stats:
        if info["source"] == "builtin":
            return "⚠️ 建议禁用（零调用内置技能）"
        elif info["source"] in ("hub", "skills.sh", "skillhub"):
            return "🗑️ 建议卸载（零调用）"
        elif info["source"] == "agent-created":
            return "⏸️ Agent创建，零调用但可保留（首次方）"
        elif info["source"] == "local":
            return "🗑️ 建议删除（零调用）"
        elif info["source"] == "external":
            return "⚠️ 外部共享，确认不影响其他 Agent 后再处理"
        return ""
    days_since = (
        (datetime.now().timestamp() - stats["last_seen"]) / 86400 if stats["last_seen"] else 999
    )
    if days_since <= 3:
        return "✅ 近 3 天使用中"
    elif days_since <= 7:
        return "✅ 近一周使用过"
    elif days_since <= 30:
        return "🟡 近一个月使用过"
    return "🟠 历史使用过"


def generate_xlsx(skills: dict[str, dict], call_stats: dict[str, dict]) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "技能审计"

    headers = ["技能名称", "来源", "启用状态", "描述（中文）", "Curator活跃度", "被合并到", "审计建议", "我的决策"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for i, w in enumerate([25, 14, 10, 60, 20, 20, 30, 12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    source_priority = {"builtin": 0, "hub": 1, "skills.sh": 1, "skillhub": 1, "agent-created": 2, "local": 3, "external": 4}
    all_items = []
    for name, info in skills.items():
        stat = call_stats.get(name)
        enabled = "禁用" if info.get("disabled") else "启用"
        suggestion = get_suggestion(name, info, stat)
        if info["source"] == "builtin":
            decision_default = "禁用" if info.get("disabled") else "启用"
        else:
            decision_default = "保留"

        # Curator 活跃度格式化
        curator_act = info.get("curator_activity", {})
        if curator_act:
            act_str = f"activity={curator_act.get('activity','?')} use={curator_act.get('use','?')} patches={curator_act.get('patches','?')} last={curator_act.get('last_activity','?')}"
        else:
            act_str = ""

        # 合并列
        merged_into = info.get("consolidated_into", "")

        all_items.append(
            {
                "name": name,
                "source": format_source(info["source"]),
                "enabled": enabled,
                "description": info.get("description", ""),
                "curator_activity": act_str,
                "consolidated_into": merged_into,
                "suggestion": suggestion,
                "disabled": info.get("disabled", False),
                "archived": info.get("archived", False),
                "source_raw": info["source"],
                "decision_default": decision_default,
            }
        )
    # Sort: disabled last, then by source priority, then archive status, then name
    all_items.sort(
        key=lambda x: (
            1 if x["disabled"] else 0,
            1 if x.get("archived") else 0,
            source_priority.get(x["source_raw"], 99),
            x["name"],
        )
    )

    fill_disabled = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    fill_delete = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
    fill_disable = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
    fill_active = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    fill_archived = PatternFill(start_color="E8E0F0", end_color="E8E0F0", fill_type="solid")

    for row_idx, item in enumerate(all_items, 2):
        values = [
            item["name"],
            item["source"],
            item["enabled"],
            item["description"],
            item["curator_activity"],
            item["consolidated_into"],
            item["suggestion"],
            item["decision_default"],
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if item["disabled"]:
                cell.fill = fill_disabled
            elif item.get("archived"):
                cell.fill = fill_archived
            elif "建议删除" in item["suggestion"] or "建议卸载" in item["suggestion"]:
                cell.fill = fill_delete
            elif "建议禁用" in item["suggestion"]:
                cell.fill = fill_disable
            elif "使用" in item["suggestion"]:
                cell.fill = fill_active

    dv_builtin = DataValidation(type="list", formula1='"启用,禁用"', allow_blank=True)
    dv_builtin.error = "请选择「启用」或「禁用」"
    dv_builtin.errorTitle = "无效输入"
    ws.add_data_validation(dv_builtin)

    dv_other = DataValidation(type="list", formula1='"保留,删除"', allow_blank=True)
    dv_other.error = "请选择「保留」或「删除」"
    dv_other.errorTitle = "无效输入"
    ws.add_data_validation(dv_other)

    for row_idx, item in enumerate(all_items, 2):
        if item["source_raw"] == "builtin":
            dv_builtin.add(f"H{row_idx}")
        else:
            dv_other.add(f"H{row_idx}")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{len(all_items) + 1}"

    output_path = str(Path.cwd() / "技能审计报告.xlsx")
    wb.save(output_path)
    return output_path


# ── 读取决策 ─────────────────────────────────────────────────────────────────
def read_decisions(xlsx_path: str) -> list[dict]:
    wb = load_workbook(xlsx_path)
    ws = wb.active
    changes = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name, source, enabled, _desc, _curator_act, _merged, _suggestion, decision = row
        if not name:
            continue
        decision = str(decision).strip() if decision else ""
        is_builtin = source == "内置"
        default = (
            "禁用" if (is_builtin and enabled == "禁用") else ("启用" if is_builtin else "保留")
        )
        if decision != default and decision:
            changes.append(
                {
                    "name": name,
                    "source": source,
                    "enabled": enabled,
                    "decision": decision,
                    "is_builtin": is_builtin,
                }
            )
    return changes


def print_changes(changes: list[dict]):
    delete_skills = [c for c in changes if c["decision"] == "删除"]
    disable_skills = [c for c in changes if c["decision"] == "禁用" and c["is_builtin"]]
    enable_skills = [c for c in changes if c["decision"] == "启用" and c["is_builtin"]]

    if not changes:
        print("无变更。所有技能决策保持默认值。")
        return False

    print("=" * 50)
    print("技能审计 — 待执行变更")
    print("=" * 50)
    if delete_skills:
        print(f"\n🗑️  删除 {len(delete_skills)} 个本地技能:")
        for c in delete_skills:
            print(f"     {c['name']}")
    if disable_skills:
        print(f"\n⚠️  禁用 {len(disable_skills)} 个内置技能:")
        for c in disable_skills:
            print(f"     {c['name']}")
    if enable_skills:
        print(f"\n✅  启用 {len(enable_skills)} 个内置技能:")
        for c in enable_skills:
            print(f"     {c['name']}")
    print("\n" + "=" * 50)
    return True


def apply_changes(changes: list[dict], skills: dict[str, dict]):
    delete_skills = [c["name"] for c in changes if c["decision"] == "删除"]
    disable_skills = [c["name"] for c in changes if c["decision"] == "禁用" and c["is_builtin"]]
    enable_skills = [c["name"] for c in changes if c["decision"] == "启用" and c["is_builtin"]]

    # Backup config.yaml
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_config = CONFIG_PATH.with_name(f"config.yaml.bak.{ts}")
    shutil.copy2(CONFIG_PATH, backup_config)
    print(f"\n✅ config.yaml 已备份: {backup_config}")

    # Backup skill directories
    backup_dir = SKILLS_DIR / ".audit-backups" / f"cleanup-{ts}"
    if delete_skills:
        backup_dir.mkdir(parents=True, exist_ok=True)
        for name in delete_skills:
            info = skills.get(name, {})
            dir_path = info.get("dir", "")
            if dir_path and Path(dir_path).exists():
                shutil.copytree(dir_path, backup_dir / name, dirs_exist_ok=True)
                shutil.rmtree(dir_path)
                print(f"  🗑️  已删除: {name}")
            else:
                print(f"  ⚪ 目录不存在: {name}")

    # Update config.yaml disabled list
    with open(CONFIG_PATH) as f:
        config = yaml.safe_load(f) or {}
    config.setdefault("skills", {})
    existing_disabled = set(config["skills"].get("disabled", []))
    new_disabled = (existing_disabled | set(disable_skills)) - set(enable_skills)
    if new_disabled != existing_disabled:
        config["skills"]["disabled"] = sorted(new_disabled)
        with open(CONFIG_PATH, "w") as f:
            yaml.dump(config, f, allow_unicode=True, default_flow_style=False)
        if disable_skills:
            print(f"  ⚠️  已禁用: {', '.join(disable_skills)}")
        if enable_skills:
            print(f"  ✅  已启用: {', '.join(enable_skills)}")

    print(f"\n✅ 执行完成。备份目录: {backup_dir if delete_skills else '（无）'}")


# ── 主函数 ───────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="Hermes Agent 技能审计工具")
    parser.add_argument("--apply", action="store_true", help="从 XLSX 读取决策并执行清理")
    parser.add_argument(
        "--xlsx", type=str, default="技能审计报告.xlsx", help="XLSX 文件路径（默认当前目录）"
    )
    args = parser.parse_args()

    if args.apply:
        xlsx_path = Path(args.xlsx)
        if not xlsx_path.exists():
            print(f"❌ 未找到 XLSX 文件: {xlsx_path}")
            print("   请先生成 XLSX: uv run audit.py")
            return

        print("📖 读取决策...")
        changes = read_decisions(str(xlsx_path))
        should_proceed = print_changes(changes)
        if not should_proceed:
            return

        print("继续执行前请确认 [y/N]: ", end="", flush=True)
        try:
            confirm = input().strip().lower()
        except (EOFError, KeyboardInterrupt):
            confirm = "n"
        if confirm != "y":
            print("已取消，未执行任何操作。")
            return

        print("\n🔧 正在执行...")
        skills = scan_all_skills()
        apply_changes(changes, skills)
        return

    # Default: scan + generate XLSX
    print("📂 扫描技能...")
    skills = scan_all_skills()
    call_stats = count_skill_calls()

    active = sum(1 for name, s in skills.items() if call_stats.get(name) and not s["disabled"])
    zero = sum(1 for name, s in skills.items() if not call_stats.get(name) and not s["disabled"])
    disabled_count = sum(1 for s in skills.values() if s["disabled"])

    # Curator 统计
    agent_created_count = sum(1 for s in skills.values() if s.get("source") == "agent-created")
    archived_count = sum(1 for s in skills.values() if s.get("archived"))
    consolidated_count = sum(1 for s in skills.values() if s.get("consolidated_into"))

    print(
        f"   共 {len(skills)} 个技能 | {active} 个在用 | {zero} 个零调用 | {disabled_count} 个已禁用"
    )
    if agent_created_count:
        print(f"   ├ Curator: {agent_created_count} 个 Agent创建 (首次方)")
    if consolidated_count:
        print(f"   ├ Curator: {consolidated_count} 个已合并到 umbrella")
    if archived_count:
        print(f"   └ Curator: {archived_count} 个已归档")

    print("📊 生成 XLSX...")
    out = generate_xlsx(skills, call_stats)
    print(f"   ✅ {out}")
    print("\n下一步：在 Excel 中打开并填写「我的决策」列，然后运行:")
    print("   uv run audit.py --apply")


if __name__ == "__main__":
    main()
